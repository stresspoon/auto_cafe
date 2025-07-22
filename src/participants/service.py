"""참여자 리스트 및 주차 정보 로딩 서비스."""

import json
import csv
from pathlib import Path
from typing import List, Dict, Optional, Union
from openpyxl import load_workbook

from ..core.logger import get_logger, log_execution_time
from ..core.exceptions import ConfigurationError
from .models import Participant, ChallengeInfo


class ParticipantLoaderService:
    """참여자 리스트와 챌린지 정보를 파일에서 로드하는 서비스."""
    
    def __init__(self, file_path: Optional[str] = None) -> None:
        """파일 경로로 서비스 초기화."""
        self._logger = get_logger(__name__)
        self._file_path = Path(file_path) if file_path else None
        self._participants: List[Participant] = []
        self._challenge_info: Optional[ChallengeInfo] = None
    
    @log_execution_time
    def load_participants(self, file_path: Optional[str] = None) -> List[Participant]:
        """파일에서 참여자 리스트를 로드."""
        path = Path(file_path) if file_path else self._file_path
        
        if not path:
            raise ConfigurationError("참여자 리스트 파일 경로가 지정되지 않았습니다")
        
        if not path.exists():
            raise ConfigurationError(f"참여자 리스트 파일을 찾을 수 없습니다: {path}")
        
        # 파일 확장자에 따라 적절한 로더 사용
        file_extension = path.suffix.lower()
        
        if file_extension == '.json':
            self._participants = self._load_from_json(path)
        elif file_extension == '.csv':
            self._participants = self._load_from_csv(path)
        elif file_extension in ['.xlsx', '.xls']:
            self._participants = self._load_from_excel(path)
        else:
            raise ConfigurationError(f"지원하지 않는 파일 형식입니다: {file_extension}")
        
        self._logger.info(f"{len(self._participants)}명의 참여자를 로드했습니다")
        return self._participants
    
    def _load_from_json(self, file_path: Path) -> List[Participant]:
        """JSON 파일에서 참여자 리스트 로드."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 챌린지 정보 로드
            if 'challenge_info' in data:
                self._challenge_info = ChallengeInfo(**data['challenge_info'])
            
            # 참여자 리스트 로드
            participants = []
            for p_data in data.get('participants', []):
                participant = Participant(
                    name=p_data['name'],
                    nickname=p_data['nickname'],
                    email=p_data.get('email'),
                    active=p_data.get('active', True)
                )
                participants.append(participant)
            
            return participants
            
        except Exception as e:
            raise ConfigurationError(f"JSON 파일 로드 중 오류: {str(e)}")
    
    def _load_from_csv(self, file_path: Path) -> List[Participant]:
        """CSV 파일에서 참여자 리스트 로드."""
        try:
            participants = []
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    participant = Participant(
                        name=row.get('name', row.get('이름', '')),
                        nickname=row.get('nickname', row.get('닉네임', '')),
                        email=row.get('email', row.get('이메일', '')),
                        active=row.get('active', row.get('활성', 'true')).lower() == 'true'
                    )
                    participants.append(participant)
            
            return participants
            
        except Exception as e:
            raise ConfigurationError(f"CSV 파일 로드 중 오류: {str(e)}")
    
    def _load_from_excel(self, file_path: Path) -> List[Participant]:
        """Excel 파일에서 참여자 리스트 로드."""
        try:
            participants = []
            wb = load_workbook(file_path, read_only=True)
            
            # 첫 번째 시트 사용
            ws = wb.active
            
            # 헤더 행 찾기
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(cell).lower() if cell else '' for cell in row]
                break
            
            # 헤더 인덱스 찾기
            name_idx = next((i for i, h in enumerate(headers) if h in ['name', '이름']), None)
            nickname_idx = next((i for i, h in enumerate(headers) if h in ['nickname', '닉네임']), None)
            email_idx = next((i for i, h in enumerate(headers) if h in ['email', '이메일']), None)
            active_idx = next((i for i, h in enumerate(headers) if h in ['active', '활성']), None)
            
            if name_idx is None or nickname_idx is None:
                raise ConfigurationError("Excel 파일에 필수 컬럼(이름, 닉네임)이 없습니다")
            
            # 데이터 행 읽기
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[name_idx] or not row[nickname_idx]:
                    continue
                
                participant = Participant(
                    name=str(row[name_idx]),
                    nickname=str(row[nickname_idx]),
                    email=str(row[email_idx]) if email_idx is not None and row[email_idx] else None,
                    active=bool(row[active_idx]) if active_idx is not None and row[active_idx] is not None else True
                )
                participants.append(participant)
            
            wb.close()
            return participants
            
        except Exception as e:
            raise ConfigurationError(f"Excel 파일 로드 중 오류: {str(e)}")
    
    @log_execution_time
    def load_challenge_info(self, file_path: Optional[str] = None) -> Optional[ChallengeInfo]:
        """챌린지 정보를 로드."""
        if self._challenge_info:
            return self._challenge_info
        
        # JSON 파일에서만 챌린지 정보를 로드 가능
        path = Path(file_path) if file_path else self._file_path
        
        if path and path.suffix.lower() == '.json':
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if 'challenge_info' in data:
                    self._challenge_info = ChallengeInfo(**data['challenge_info'])
                    self._logger.info(f"챌린지 정보 로드: {self._challenge_info.name}")
                
            except Exception as e:
                self._logger.warning(f"챌린지 정보 로드 실패: {str(e)}")
        
        return self._challenge_info
    
    def get_participants(self) -> List[Participant]:
        """로드된 참여자 리스트 반환."""
        return self._participants
    
    def get_active_participants(self) -> List[Participant]:
        """활성 참여자만 반환."""
        return [p for p in self._participants if p.active]
    
    def get_participant_by_nickname(self, nickname: str) -> Optional[Participant]:
        """닉네임으로 참여자 검색."""
        normalized_nickname = nickname.strip()
        for participant in self._participants:
            if participant.nickname == normalized_nickname:
                return participant
        return None
    
    def get_challenge_info(self) -> Optional[ChallengeInfo]:
        """챌린지 정보 반환."""
        return self._challenge_info
    
    def to_dict(self) -> Dict[str, Union[List[dict], dict]]:
        """전체 데이터를 딕셔너리로 변환."""
        result = {
            'participants': [p.to_dict() for p in self._participants]
        }
        
        if self._challenge_info:
            result['challenge_info'] = self._challenge_info.to_dict()
        
        return result